import pysam
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

# Variant Parsing
def parse_vcf(vcf_file):
    variants = []
    with pysam.VariantFile(vcf_file) as vcf:
        for i, variant in enumerate(vcf):
            if i >= 10:  # Stop parsing after 5 rows
                break
            chromosome = variant.chrom
            position = variant.pos
            ref_allele = variant.ref
            alt_allele = ",".join(map(str, variant.alts))
            quality = variant.qual
            info = variant.info
            filter_value = variant.filter.keys()[0] if variant.filter else "PASS"  # Get filter value or default to "PASS"
            format_fields = variant.format
            
            variants.append({
                "Chromosome": chromosome,
                "Position": position,
                "Ref Allele": ref_allele,
                "Alt Allele": alt_allele,
                "Quality": quality,
                "Annotations": info,
                "FILTER": filter_value,
                "FORMAT": format_fields,
            })
    return variants

# Variant Classification (Placeholder function)
def classify_variants(variants):
    for variant in variants:
        variant['Classification'] = 'Unclassified'
    return variants

# Report Generation
def generate_report(variants, output_file):
    wb = Workbook()
    ws = wb.active

    title = "Clinical Report"
    ws.append([title])

    column_headers = list(variants[0].keys())
    ws.append(column_headers)

    for variant in variants:
        ws.append([str(variant[key]) for key in column_headers])

    # Set font and alignment for title
    title_cell = ws['A1']
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal='center')

    # Set font and alignment for column headers
    for col in ws.iter_cols(min_row=2, max_row=2, min_col=1, max_col=len(column_headers)):
        for cell in col:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')

    wb.save(output_file)

if __name__ == "__main__":
    vcf_file = "/Users/mycetism/Desktop/normal_sample.deepvariant.vcf"
    output_file = "/Users/mycetism/Desktop/Python projects/Bioinformatics/clinical_report.xlsx"

    # Variant Parsing
    variants = parse_vcf(vcf_file)

    # Variant Classification
    variants = classify_variants(variants)

    # Report Generation
    generate_report(variants, output_file)

    print("Clinical report generated successfully.")
